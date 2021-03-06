import math
import openpyxl as xls
import numpy as np

def F(x):
    '''
        Функция, которую исслудуем по таблице.
    '''
    return (math.cos(x) - x)

def parse_table():
    '''
        Загрузка таблицы в программу.
    '''
    pos = 3
    points = xls.load_workbook("points.xlsx").active
    table = []
    while points.cell(row = pos, column = 1).value is not None:
        table.append([float(points.cell(row = pos, column = 1).value), \
                      float(points.cell(row = pos, column = 2).value), \
                      float(points.cell(row = pos, column = 3).value)])
        pos += 1
    # table.sort()
    count = len(table) - 1
    arr_x = []
    arr_y = []
    for i in range(len(table)):
        arr_x.append(table[i][0])
        arr_y.append(table[i][1])
    return table, arr_x, arr_y, count

def input_x():
    '''
        Ввод аргумента. (в случае ошибки дается еще попытка)
    '''
    print("Enter X: ")
    flag = 0
    x = 0
    while flag == 0:
        x = 1.5
        try:
            val = float(x)
            flag = 1
        except ValueError:
            print("Some error! Try again")
    return float(x)

def find_x0_xn(data, power, arg):
    '''
        Нахождение начального и конечного индекса в таблице (x0 и xn).
    '''
    index_x = 0

    while arg > data[index_x][0]:
        index_x += 1
    index_x0 = index_x - power // 2 - 1
    index_xn = index_x + (power // 2) + (power % 2) - 1
    if index_xn > len(data) - 1:
        index_x0 -= index_xn - len(data) + 1
        index_xn = len(data) - 1
    elif index_x0 < 0:
        index_xn += -index_x0
        index_x0 = 0
    return index_x0, index_xn

def div_diff(x, y, node):
    '''
        Расчет разделенных разниц для полинома Ньютона
    '''
    pol = []
    for i in range(node):
        pol.append([0] * (node + 1))
    for i in range(node):
        pol[i][0], pol[i][1] = x[i], y[i]
    i = 2
    new_node = node - 1
    while i < (node + 1):
        j = 0
        while j < new_node:
            pol[j][i] = round((pol[j + 1][i - 1] - pol[j][i - 1]) \
                 / (pol[i - 1][0] - pol[0][0]), 5)
            j += 1
        i += 1
        new_node -= 1
    return pol

def polinom_n(x, y, node, arg):
    '''
        Расчет значение функции от заданного аргумента.
        Полином Ньютона.
    '''
    pol = div_diff(x, y, node)
    y = pol[0][1]
    i = 2
    while i < node + 1:
        j, p = 0, 1
        while j < i - 1:
            p *= (arg - pol[j][0])
            j += 1
        y += pol[0][i] * p
        i += 1
    return y

def hermite_interpolate(data, node, arg, coords_x):
    '''
        Расчет таблицы для полинома Эрмита.
    '''
    # Поиск нужных начала и конца отрезка х
    x0, xn = find_x0_xn(data, node // 2, arg)
    data = data[x0 : xn + 1]
    pol = []
    for i in range(2 * len(data)):
        pol.append([0] * (2 * node + 3))
    i = 0
    for j in range(len(data)):
        pol[i][0], pol[i][1], pol[i][2] = data[j][0], data[j][1], data[j][2]
        i += 1
        pol[i][0], pol[i][1] = data[j][0], data[j][1]
        i += 1
    i = 2
    # Заполнения таблицы, как для полнома Ньютона
    for j in range(len(pol) - 1):
        if j % 2 == 1:
            pol[j][i] = (pol[j][1] - pol[j + 1][1]) \
                 / (pol[j][0] - pol[j + 1][0])
    i = 3
    new_node = node - 2
    while i < len(pol):
        j = 0
        while j < new_node:
            pol[j][i] = round((pol[j + 1][i - 1] - pol[j][i - 1]) \
                 / (pol[i - 1][0] - pol[0][0]), 5)
            j += 1
        i += 1
        new_node -= 1
    return pol

def polynom_h(pol, node, arg):
    '''
        Расчет значение функции от заданного аргумента.
        Полином Эрмита.
    '''
    y = pol[0][1]
    i = 2
    while i < node + 2:
        j = 0
        p = 1
        while j < i - 1:
            p *= (arg - pol[j][0])
            j += 1
        y += pol[0][i] * p
        i += 1
    return y

def main():
    data, coords_x, coords_y, count = parse_table()
    x = input_x()
    arr_n = [3]
    # arr_n = [0, 1, 2, 3, 4]

    print("\nИнтерполяция с помощью полинома Ньютона и Эрмита\n",
          "| n |   x | п. Ньютона | п. Эрмита |")
    for n in arr_n:
        print(" | {} | {} |".format(n, x), end="")
        flag = False
        # Проверка на наличие данного аргумента в таблице
        for i in range(0, len(data)):
            if x == data[i][0]:
                print("     {}    |".format(round(data[i][1], 5)), end="")
                flag = True
        # В случае, если агрумента в таблице нет
        if not flag:
            # Полином Ньютона
            x0, xn = find_x0_xn(data, n, x)
            ax = coords_x[x0 : xn + 1]
            ay = coords_y[x0 : xn + 1]
            if len(ax):
                my_root = polinom_n(ax, ay, n + 1, x)
                print("   {}  |".format(round(my_root, 5)), end="")

            # # Полином Эрмита
            pol = hermite_interpolate(data, n, x, coords_x)
            my_root2 = polynom_h(pol, n, x)
            print("  {}  |".format(round(my_root2, 5)), end="\n")

    print("\nОбратная инерполяция с помощью полинома Ньютона\n",
          "| n |  x  |  Корень  |")
    for n in arr_n:
        print(" | {} |  {}  |".format(n, 0), end="")
        flag = False
        # Проверка на наличие данного аргумента в таблице
        for i in range(0, len(data)):
            if 0 == data[i][1]:
                print("  {} |".format(round(data[i][1], 5)), end="")
                flag = True
        # В случае, если агрумента в таблице нет
        if not flag:
            n_data = []
            for i in range(len(data)):
                n_data.append([data[i][1], data[i][0], data[i][2]])
            n_data.sort()
            coords_y.clear()
            coords_x.clear()
            for i in range(len(n_data)):
                coords_x.append(n_data[i][0])
                coords_y.append(n_data[i][1])
            x0, xn = find_x0_xn(n_data, n, 0)
            ax = coords_x[x0 : xn + 1]
            ay = coords_y[x0 : xn + 1]
            if len(ax):
                my_root = polinom_n(ax, ay, n + 1, 0)
                print("  {} |".format(round(my_root, 5)))

if __name__ == "__main__":
    main()
